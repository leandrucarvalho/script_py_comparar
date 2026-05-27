"""
Comparador de Concluintes BFD
==============================
Compara os nomes da planilha Excel com as pastas locais (Drive baixado).

Uso:
    python comparar_concluintes.py

Configure os caminhos abaixo antes de rodar.
"""

import os
import re
import unicodedata
import openpyxl
from difflib import SequenceMatcher
from openpyxl.styles.fills import PatternFill

# openpyxl bug: some Excel files include an 'extLst' XML attribute in fill
# styles that PatternFill.__init__ doesn't accept. Patch it to ignore extras.
_orig_pf_init = PatternFill.__init__
def _pf_init(self, *args, **kwargs):
    kwargs.pop("extLst", None)
    _orig_pf_init(self, *args, **kwargs)
PatternFill.__init__ = _pf_init

# ─────────────────────────────────────────────
# CONFIGURAÇÃO — ajuste esses caminhos
# ─────────────────────────────────────────────

EXCEL_PATH = r"C:\Users\leo-s\Downloads\BFD_Lista de Concluintes_Ciclo1_por Estado_por Turma.xlsx"
SHEET_NAME = "Leandro"
ESTADOS    = ["Rio de Janeiro"]


# Pasta raiz onde estão as subpastas de cada estado > turma > aluno
# Exemplo de estrutura esperada:
#   EVIDENCIAS/
#     GOIAS/
#       T01GOC1/
#         daivid_alves_de_souza/
#         felipe_mendes_alves/
#     MARANHAO/
#       T01MAC1/
#         ...
DRIVE_ROOT = r"C:\Users\leo-s\Downloads\Rio de Janeiro"

# ─────────────────────────────────────────────
# FUNÇÕES AUXILIARES
# ─────────────────────────────────────────────

def normalizar(texto):
    """Remove acentos, pontuação, espaços extras e coloca em minúsculas."""
    if not texto:
        return ""
    texto = str(texto).lower().strip()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"[^a-z0-9 ]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def _palavras_parecidas(a, b, threshold=0.82):
    return SequenceMatcher(None, a, b).ratio() >= threshold


def nomes_em_comum(nome_a, nome_b, min_palavras=2):
    """
    Retorna True se os dois nomes compartilham pelo menos `min_palavras`
    palavras significativas (ignora partículas como de/da/do/dos/das/e).
    Aceita pequenos erros de digitação via similaridade de texto.
    """
    particulas = {"de", "da", "do", "dos", "das", "e", "a", "o"}
    pa = [w for w in normalizar(nome_a).split() if w not in particulas]
    pb = [w for w in normalizar(nome_b).split() if w not in particulas]

    usados = set()
    matches = 0
    for wa in pa:
        for i, wb in enumerate(pb):
            if i not in usados and _palavras_parecidas(wa, wb):
                matches += 1
                usados.add(i)
                break
    return matches >= min_palavras


# ─────────────────────────────────────────────
# LEITURA DO EXCEL
# ─────────────────────────────────────────────

def ler_planilha(excel_path, sheet_name, estados):
    """Retorna dict: {turma: [lista de nomes normalizados]}"""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba '{sheet_name}' não encontrada. Abas disponíveis: {wb.sheetnames}")

    ws = wb[sheet_name]
    dados = {}  # turma -> [nomes]

    estados_norm = [normalizar(e) for e in estados]

    for row in ws.iter_rows(values_only=True):
        estado, turma, nome = row[0], row[1], row[2]
        if not estado or not turma or not nome:
            continue
        if normalizar(str(estado)) not in estados_norm:
            continue
        turma = str(turma).strip()
        nome  = str(nome).strip()
        if turma not in dados:
            dados[turma] = []
        dados[turma].append(nome)

    wb.close()
    return dados


# ─────────────────────────────────────────────
# LEITURA DAS PASTAS LOCAIS
# ─────────────────────────────────────────────

def normalizar_turma(nome):
    """
    Padroniza o nome da turma removendo prefixo 'Turma' se houver.
    Ex: 'Turma03PEC1' -> 'T03PEC1'
    """
    nome = nome.strip()
    if nome.lower().startswith("turma"):
        nome = "T" + nome[5:]
    return nome


def ler_pastas(drive_root):
    """
    Percorre drive_root até 1 nível:
      turma/ > aluno/
    Retorna dict: {turma_normalizada: [nomes das subpastas de alunos]}
    """
    pastas = {}

    if not os.path.isdir(drive_root):
        raise FileNotFoundError(f"Pasta não encontrada: {drive_root}")

    for turma_dir in os.scandir(drive_root):
        if not turma_dir.is_dir():
            continue
        turma = normalizar_turma(turma_dir.name)
        alunos = [
            a.name for a in os.scandir(turma_dir.path) if a.is_dir()
        ]
        pastas[turma] = alunos

    return pastas


# ─────────────────────────────────────────────
# COMPARAÇÃO
# ─────────────────────────────────────────────

def comparar(planilha, drive):
    extras   = {}   # no drive mas não na planilha
    faltando = {}   # na planilha mas não no drive

    todas_turmas = sorted(set(list(planilha.keys()) + list(drive.keys())))

    for turma in todas_turmas:
        plan_nomes  = planilha.get(turma, [])
        drive_nomes = drive.get(turma, [])

        # Extras no drive
        ex = []
        for d in drive_nomes:
            encontrado = any(nomes_em_comum(d, p) for p in plan_nomes)
            if not encontrado:
                ex.append(d)
        if ex:
            extras[turma] = ex

        # Faltando no drive
        fa = []
        for p in plan_nomes:
            encontrado = any(nomes_em_comum(p, d) for d in drive_nomes)
            if not encontrado:
                fa.append(p)
        if fa:
            faltando[turma] = fa

    return extras, faltando


# ─────────────────────────────────────────────
# RELATÓRIO
# ─────────────────────────────────────────────

def imprimir_relatorio(extras, faltando):
    SEP = "=" * 65

    print(SEP)
    print("❌  PASTAS NO DRIVE QUE NÃO ESTÃO NA PLANILHA")
    print("    (candidatas a remoção)")
    print(SEP)
    if extras:
        total = 0
        for turma, nomes in sorted(extras.items()):
            print(f"\n  {turma}:")
            for n in nomes:
                print(f"    - {n}")
                total += 1
        print(f"\n  Total: {total} pasta(s) extra(s)")
    else:
        print("  Nenhuma pasta extra encontrada. ✅")

    print()
    print(SEP)
    print("⚠️   ALUNOS NA PLANILHA SEM PASTA NO DRIVE")
    print("    (evidências ausentes)")
    print(SEP)
    if faltando:
        total = 0
        for turma, nomes in sorted(faltando.items()):
            print(f"\n  {turma}:")
            for n in nomes:
                print(f"    - {n}")
                total += 1
        print(f"\n  Total: {total} aluno(s) sem pasta")
    else:
        print("  Todos os alunos têm pasta no Drive. ✅")

    print()


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print(f"\nLendo planilha: {EXCEL_PATH}")
    planilha = ler_planilha(EXCEL_PATH, SHEET_NAME, ESTADOS)
    print(f"  {sum(len(v) for v in planilha.values())} alunos em {len(planilha)} turmas\n")

    print(f"Lendo pastas: {DRIVE_ROOT}")
    drive = ler_pastas(DRIVE_ROOT)
    print(f"  {sum(len(v) for v in drive.values())} pastas em {len(drive)} turmas\n")

    extras, faltando = comparar(planilha, drive)
    imprimir_relatorio(extras, faltando)